import re
from typing import Optional
from openpyxl.worksheet.worksheet import Worksheet

def parse_curriculum_metadata(sheet: Worksheet) -> dict:
    data = {
        'major_type': None,
        'major_duration': None,  # integer (years)
        'education_type': None,
        'major_code': None,
        'major_title': None,
        'mandatory_courses': [],  # List of dicts: {code, title}
    }
    ta_lim_yonalishi_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        for col_idx, cell in enumerate(row):
            if cell and isinstance(cell, str):
                normalized = (
                    cell.replace("‘", "'")
                        .replace("’", "'")
                        .replace("ʼ", "'")
                        .replace("`", "'")
                        .replace("–", "-")
                        .replace("—", "-")
                        .replace("−", "-")
                )
                text = normalized.lower()
                if 'akademik daraja' in text and data['major_type'] is None:
                    parts = re.split(r'-|:', normalized)
                    if len(parts) > 1:
                        data['major_type'] = parts[1].strip()
                if "o'qish muddati" in text and data['major_duration'] is None:
                    parts = re.split(r'-|:', normalized)
                    if len(parts) > 1:
                        # Extract integer from e.g. '4 yil'
                        match = re.search(r'(\d+)', parts[1])
                        data['major_duration'] = int(match.group(1)) if match else None
                if "ta'lim shakli" in text and data['education_type'] is None:
                    parts = re.split(r'-|:', normalized)
                    if len(parts) > 1:
                        data['education_type'] = parts[1].strip()
                if "ta'lim yo'nalishi" in text and ta_lim_yonalishi_row is None:
                    ta_lim_yonalishi_row = row_idx
    if ta_lim_yonalishi_row:
        next_row = list(sheet.iter_rows(min_row=ta_lim_yonalishi_row+1, max_row=ta_lim_yonalishi_row+1, values_only=True))
        if next_row:
            for cell in next_row[0]:
                if cell and isinstance(cell, str):
                    normalized_cell = (
                        cell.replace("–", "-")
                            .replace("—", "-")
                            .replace("−", "-")
                            .replace("  ", " ")
                            .strip()
                    )
                    # Match code-title pattern
                    m = re.match(r"^(\d{5,})\s*-\s*(.+)$", normalized_cell)
                    if m:
                        data['major_code'] = m.group(1)
                        data['major_title'] = m.group(2).strip()
                        break
    # --- Parse mandatory courses ---
    found_first_course = False
    idx_101 = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if not found_first_course:
            if row and any(str(cell).strip() == '1.01' for cell in row if cell):
                found_first_course = True
                idx_101 = next((i for i, cell in enumerate(row) if str(cell).strip() == '1.01'), None)
                if idx_101 is not None:
                    code = str(get_merged_cell_value(sheet, row_idx + 1, idx_101 + 1)).strip() if get_merged_cell_value(sheet, row_idx + 1, idx_101 + 1) else None
                    title = str(get_merged_cell_value(sheet, row_idx + 1, idx_101 + 2)).strip() if get_merged_cell_value(sheet, row_idx + 1, idx_101 + 2) else None
                    hours = [
                        get_merged_cell_value(sheet, row_idx + 1, idx_101 + 3),
                        get_merged_cell_value(sheet, row_idx + 1, idx_101 + 4),
                        get_merged_cell_value(sheet, row_idx + 1, idx_101 + 5),
                        get_merged_cell_value(sheet, row_idx + 1, idx_101 + 6),
                    ]
                    if code and title:
                        data['mandatory_courses'].append({'code': code, 'title': title, 'hours': hours})
                continue
        if found_first_course:
            if not row or not any(row):
                break  # Stop at first empty row
            code = str(get_merged_cell_value(sheet, row_idx + 1, idx_101 + 1)).strip() if get_merged_cell_value(sheet, row_idx + 1, idx_101 + 1) else None
            title = str(get_merged_cell_value(sheet, row_idx + 1, idx_101 + 2)).strip() if get_merged_cell_value(sheet, row_idx + 1, idx_101 + 2) else None
            hours = [
                get_merged_cell_value(sheet, row_idx + 1, idx_101 + 3),
                get_merged_cell_value(sheet, row_idx + 1, idx_101 + 4),
                get_merged_cell_value(sheet, row_idx + 1, idx_101 + 5),
                get_merged_cell_value(sheet, row_idx + 1, idx_101 + 6),
            ]
            if code and title:
                data['mandatory_courses'].append({'code': code, 'title': title, 'hours': hours})
            else:
                break  # Stop if code or title is missing
    return data

def get_merged_cell_value(sheet: Worksheet, row_idx: int, col_idx: int):
    cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
    for merged_range in sheet.merged_cells.ranges:
        if (cell.coordinate in merged_range):
            # Get the top-left cell of the merged range
            tl_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            return tl_cell.value
    return cell.value 